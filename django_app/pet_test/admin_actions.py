# -*- coding: utf-8 -*-
import datetime
import logging

from django.contrib import admin
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse
from openpyxl import load_workbook
import io

from .admin import ImportedData

logger = logging.getLogger('')


class UploadedFilesAdmin(admin.ModelAdmin):
    actions = ['process_xls_file']

    def process_xls_file(self, request, queryset):
        """
        Procesa un archivo excel y pone los datos en el modelo
        """
        if not request.user.is_staff:
            raise PermissionDenied

        opts = self.model._meta
        field_name = opts.fields[1]
        for obj in queryset:
            filename = getattr(obj, field_name.name)

            # Abrir el excel
            wb = load_workbook(filename, read_only=True)
            sh = wb.active

            # obtener los nombres de las columnas
            col_names = []
            for col in range(1, sh.max_column):
                col_names.append(sh.cell(column=col, row=1).value)

            for row in range(2, sh.max_row):
                imp = ImportedData(
                    CODIGO=sh.cell(column=1, row=row).value,
                    DETALLE=sh.cell(column=2, row=row).value,
                    SUBRUBRO=sh.cell(column=3, row=row).value,
                    RUBRO=sh.cell(column=4, row=row).value,
                    MARCA_P=sh.cell(column=5, row=row).value,
                    EDAD=sh.cell(column=6, row=row).value,
                    CONSIDER=sh.cell(column=7, row=row).value,
                    GRUPO=sh.cell(column=8, row=row).value,
                    PESO=sh.cell(column=9, row=row).value,
                    MATERIAL=sh.cell(column=10, row=row).value,
                    TIPO_P=sh.cell(column=11, row=row).value,
                    USO=sh.cell(column=12, row=row).value,
                    TAMANO=sh.cell(column=13, row=row).value,
                    URLFOTO1=sh.cell(column=14, row=row).value,
                    URLFOTO2=sh.cell(column=15, row=row).value,
                    URLFOTO3=sh.cell(column=16, row=row).value,
                    PRECIO=sh.cell(column=17, row=row).value,
                    STOCK=sh.cell(column=18, row=row).value,
                    LISTAPRE=sh.cell(column=19, row=row).value,
                    TITUMELI=sh.cell(column=20, row=row).value,
                    GRUPO_RECA=sh.cell(column=21, row=row).value,
                    DESC_LAR=sh.cell(column=22, row=row).value,
                )
                """
                for col in range(1, sh.max_column):
                    value = sh.cell(column=col, row=row).value
                    logger.warning('value %s, data %s, col %s ---', value,
                                   col_names[col - 1], col)
                    f = imp._meta.get_field(col_names[col - 1])
                    f.default = value

                    logger.warning('field [%s] [%s]', f.name, f.default)
                """
                imp.save()
                logger.warning('SAVED ----------- %s ',row)

    process_xls_file.short_description = "Procesar archivo excel"


class ImportedDataAdmin(admin.ModelAdmin):
    list_display = [
        'CODIGO',
        'DETALLE',
        'SUBRUBRO',
        'RUBRO',
        'MARCA_P',
        'EDAD',
        'CONSIDER',
        'GRUPO',
        'PESO',
        'MATERIAL',
        'TIPO_P',
        'USO',
        'TAMANO',
        'URLFOTO1',
        'URLFOTO2',
        'URLFOTO3',
        'PRECIO',
        'STOCK',
        'LISTAPRE',
        'TITUMELI',
        'GRUPO_RECA',
        'DESC_LAR',
    ]


class GuestsAdmin(admin.ModelAdmin):
    actions = ['export_as_xls']

    def export_as_xls(self, request, queryset):
        """
        Generic xls export admin action.
        """
        if not request.user.is_staff:
            raise PermissionDenied
        opts = self.model._meta
        output = io.BytesIO()
        wb = Workbook(output, {'in_memory': True})
        ws0 = wb.add_worksheet()
        col = 0
        field_names = []
        # write header row
        for field in opts.fields:
            ws0.write(0, col, field.name)
            field_names.append(field.name)
            col = col + 1
        row = 1
        # Write data rows
        for obj in queryset:
            col = 0
            for field in field_names:
                #               val = unicode(getattr(obj, field)).strip()
                val = getattr(obj, field)
                if type(val) == type(datetime.datetime.now()):
                    val = val.date().strftime('%d/%m/%Y')  # replace(tzinfo=None)
                ws0.write(row, col, val)
                col = col + 1
            row = row + 1

        wb.close()
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = "attachment; filename=export.xlsx"
        return response

    export_as_xls.short_description = "Exportar como planilla excel (xlsx)"
